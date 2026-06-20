from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO, StringIO
import csv
from pathlib import Path
from typing import Any

from fastapi import UploadFile
from openpyxl import load_workbook


class UploadParsingError(ValueError):
    pass


@dataclass(frozen=True)
class ParsedUploadRow:
    row_number: int
    data: dict[str, Any]


@dataclass(frozen=True)
class ParsedUpload:
    headers: list[str]
    rows: list[ParsedUploadRow]


async def parse_case_upload_file(file: UploadFile) -> ParsedUpload:
    extension = Path(file.filename or "").suffix.lower()
    content = await file.read()

    if not content:
        raise UploadParsingError("Uploaded file is empty.")

    if extension == ".csv":
        return _parse_csv(content)

    if extension in {".xlsx", ".xlsm"}:
        return _parse_xlsx(content)

    raise UploadParsingError("Unsupported file type. Upload a CSV or XLSX file.")


def _parse_csv(content: bytes) -> ParsedUpload:
    text = _decode_csv(content)
    reader = csv.reader(StringIO(text))
    rows = list(reader)

    if not rows:
        raise UploadParsingError("Uploaded CSV does not contain a header row.")

    headers = _normalize_headers(rows[0])
    if not any(headers):
        raise UploadParsingError("Uploaded CSV header row is empty.")

    parsed_rows = [
        ParsedUploadRow(row_number=index, data=_row_to_dict(headers, row))
        for index, row in enumerate(rows[1:], start=2)
        if not _is_empty_row(row)
    ]

    return ParsedUpload(headers=headers, rows=parsed_rows)


def _parse_xlsx(content: bytes) -> ParsedUpload:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise UploadParsingError("Uploaded XLSX file could not be read.") from exc

    worksheet = workbook.active
    raw_rows = list(worksheet.iter_rows(values_only=True))

    if not raw_rows:
        raise UploadParsingError("Uploaded XLSX does not contain a header row.")

    headers = _normalize_headers(raw_rows[0])
    if not any(headers):
        raise UploadParsingError("Uploaded XLSX header row is empty.")

    parsed_rows = [
        ParsedUploadRow(row_number=index, data=_row_to_dict(headers, list(row)))
        for index, row in enumerate(raw_rows[1:], start=2)
        if not _is_empty_row(row)
    ]

    return ParsedUpload(headers=headers, rows=parsed_rows)


def _decode_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue

    raise UploadParsingError("Uploaded CSV encoding is not supported.")


def _normalize_headers(headers: list[Any] | tuple[Any, ...]) -> list[str]:
    seen: dict[str, int] = {}
    normalized_headers: list[str] = []

    for index, header in enumerate(headers, start=1):
        header_text = str(header or "").strip() or f"unnamed_column_{index}"
        count = seen.get(header_text, 0)
        seen[header_text] = count + 1
        normalized_headers.append(header_text if count == 0 else f"{header_text}_{count + 1}")

    return normalized_headers


def _row_to_dict(headers: list[str], row: list[Any] | tuple[Any, ...]) -> dict[str, Any]:
    return {
        header: row[index] if index < len(row) else None
        for index, header in enumerate(headers)
    }


def _is_empty_row(row: list[Any] | tuple[Any, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in row)
